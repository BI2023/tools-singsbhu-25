import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
input_file = "input_file.xlsx"  # File with multiple sheets
search_file = "search_file.xlsx"  # File with search criteria

# Load the search file into a DataFrame
search_df = pd.read_excel(search_file)

# Load the input file (it may contain multiple sheets)
input_file_data = pd.ExcelFile(input_file)

# Create the yellow and red fill styles
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Load the workbook of the search file to modify it
wb = load_workbook(search_file)
ws = wb.active  # Assuming search data is in the first sheet (change if needed)

# Columns from search file
sheet_column = "SheetName"  # Column with sheet names
search_column1 = "SearchColumn1"  # First search column
search_column2 = "SearchColumn2"  # Second search column

# Iterate through rows in the search file
for idx, row in search_df.iterrows():
    sheet_name = row[sheet_column]
    value1 = row[search_column1]
    value2 = row[search_column2]

    # Check if the sheet exists in the input file
    if sheet_name in input_file_data.sheet_names:
        # Load the target sheet from the input file
        df = pd.read_excel(input_file, sheet_name=sheet_name)

        # Assuming we're searching in specific columns in the target sheet
        target_column1 = "ColumnToSearch1"  # Update with correct column name
        target_column2 = "ColumnToSearch2"  # Update with correct column name

        # Check if both columns exist
        if target_column1 in df.columns and target_column2 in df.columns:
            # Check if the combination exists in the target sheet
            matched_rows = df[(df[target_column1] == value1) & (df[target_column2] == value2)]
            
            # Highlight the row yellow if found
            if not matched_rows.empty:
                ws[f"A{idx + 2}"].parent.row_dimensions[idx + 2].fill = yellow_fill  # Adjust for row number
                print(f"Row {idx + 2} highlighted yellow")
                continue

    # If not found, highlight the row red
    ws[f"A{idx + 2}"].parent.row_dimensions[idx + 2].fill = red_fill  # Adjust for row number
    print(f"Row {idx + 2} highlighted red")

# Save the modified search file
wb.save("highlighted_search_file.xlsx")
print("Search file saved with highlights.")
